
import nltk
from nltk import sent_tokenize, word_tokenize
from nltk.corpus import stopwords
from nltk.stem.porter import PorterStemmer
from nltk.stem import WordNetLemmatizer
import pandas as pd
import openpyxl
from textblob import TextBlob
from nltk import pos_tag
#----------------_______________________initializers_______________--------------------#
stop_words = set(stopwords.words('english'))
stop_words.update('I','am','the')
#---------------____________________word correction feature can be added later__________________---------------#
#corrected_words_dict={'rac':"rack"}
corrected_tokens=[]
porter_stemmer=PorterStemmer()
Word_Net_Lemmatizer=WordNetLemmatizer()
stemmed_tokens=[]
lemmatized_tokens=[]
simple_tokens=[]
noun_phrases=[]
year=set([])
nouns=set([])
swear_words=set([])
#----------------------______________Excel file datatsets_________________-----------------------------#
file='us_business_database_sample.xlsx'
Xl=openpyxl.load_workbook(file)
sheet=(Xl.get_sheet_by_name('Sheet1'))
col=1
file2='fortune1000-2012.xlsx'
Xl2=openpyxl.load_workbook(file2)
sheet2=(Xl2.get_sheet_by_name('Sheet1'))
col=1
file3='swearWords.xlsx'
Xl3=openpyxl.load_workbook(file3)
sheet3=(Xl3.get_sheet_by_name('Worksheet'))
col=2

#------------------------____________SETS______________________-------------------------------#
Universal_op=set(['add','subtract','multiply','division','percentile','modulus'])
Addition_op=set(['total','sum','add','aggregate','addition','plus'])
Difference_op=set(['difference','minus','subtract'])
Multiply_op=set(['multiply'])
Division_op=set(['divide','division'])
basic_op=set(Addition_op | Difference_op | Multiply_op | Division_op | Universal_op)
company_name=set([])
time=set(['1','2','3','4','5','6','7','8','9','0','year','date','month','time','day','quarter','quarterly','annual','annually'])
numbers=set([])
for i in range(1000):
    numbers.add(str(i))
add='add'
sub='subtract'
mul='multiply'
div='divide'
month=set(['january','february','march','april','may','june','july','august','september','october','november','december'])
day=set(['monday','tuesday','wednesday','thursday','friday','saturday','sunday'])
for i in range(1,3):
    for j in range(0,10):
        for k in range(0,10):
            for l in range(0,10):
                year.add(str(1000*i+100*j+10*k+l))
for i in range(1,74):
    stop_words.add(sheet3.cell(row=1,column=i).value)
#----------------------_______________________sentence_____________________---------------------------------#



sentence="whats the total amount of income for joe this january"


#-----------------------______________________word tokens___________________---------------------------#
word_tokens=nltk.word_tokenize(sentence.lower())
#----------------------_____________________stop_words________________---------------------------#
extracted_tokens=list(word_tokens[i] for i in range(len(word_tokens)) if word_tokens[i] not in stop_words)
#-----------------------________________corrected and lemmatized tokens____________-------------------------#
for i in range(len(extracted_tokens)):
    if extracted_tokens[i] in corrected_words_dict.keys():
        corrected_tokens.append(corrected_words_dict[extracted_tokens[i]])
        stemmed_tokens.append(porter_stemmer.stem(corrected_tokens[i]))
        lemmatized_tokens.append(Word_Net_Lemmatizer.lemmatize(corrected_tokens[i]))
    else:
        corrected_tokens.append(extracted_tokens[i])
        stemmed_tokens.append(porter_stemmer.stem(corrected_tokens[i]))
        lemmatized_tokens.append(Word_Net_Lemmatizer.lemmatize(corrected_tokens[i]))
#print(lemmatized_tokens)
#------------------------_____________________Noun_____________________________---------------------------#
pos=nltk.pos_tag(lemmatized_tokens)
nouns.update(pos[i][0] for i in range(len(lemmatized_tokens)) if pos[i][1]=='NN')
#------------------------____________________company name________________------------------------#
for i in range(2,98):
    company_name.add(sheet.cell(row=i,column=1).value)
company_name.update(['Oracle'],['Facebook'],['Apple'],['Netflix'],['Amazon'])
for i in range(7,1002):
    company_name.add(sheet2.cell(row=i,column=1).value)
#for i in range(6,1004):
#    company_name.add(sheet3.cell(row=i,column=2).value)
#-----------------____________________basic op or company name--------------------____________________#
for i in range(len(lemmatized_tokens)):
    if(lemmatized_tokens[i] in basic_op | company_name | nouns | year | month | day | numbers | time ):
        simple_tokens.append(lemmatized_tokens[i])
#--------------------------__________________operators__________________-------------------------#
for i in range(len(simple_tokens)):
        if(simple_tokens[i] in Addition_op):
                simple_tokens[i]=add
        elif(simple_tokens[i] in Difference_op):
                simple_tokens[i]=sub
        elif(simple_tokens[i] in Multiply_op):
                simple_tokens[i]=mul
        elif(simple_tokens[i] in Division_op):
                simple_tokens[i]=div
print(simple_tokens)
